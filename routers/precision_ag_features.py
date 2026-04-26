"""
Precision Ag — Scouting, Soil Samples, Prescriptions, Weather, Reports
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc
from database import get_db
from datetime import datetime, date
import json, csv, io, os, requests
import models

router = APIRouter(prefix="/api", tags=["precision-ag-features"])

# crop-type baseline yields (kg/ha) and GDD base temperatures (°F)
_CROP_BASELINES = {
    "wheat":   {"yield_kgha": 3500,  "gdd_base_f": 40, "kc": 1.0},
    "corn":    {"yield_kgha": 9000,  "gdd_base_f": 50, "kc": 1.15},
    "maize":   {"yield_kgha": 9000,  "gdd_base_f": 50, "kc": 1.15},
    "soy":     {"yield_kgha": 2800,  "gdd_base_f": 50, "kc": 1.0},
    "soybean": {"yield_kgha": 2800,  "gdd_base_f": 50, "kc": 1.0},
    "canola":  {"yield_kgha": 2000,  "gdd_base_f": 41, "kc": 1.0},
    "cotton":  {"yield_kgha": 1800,  "gdd_base_f": 60, "kc": 1.15},
    "rice":    {"yield_kgha": 6000,  "gdd_base_f": 50, "kc": 1.2},
    "barley":  {"yield_kgha": 3200,  "gdd_base_f": 40, "kc": 1.0},
    "oats":    {"yield_kgha": 2800,  "gdd_base_f": 40, "kc": 1.0},
    "default": {"yield_kgha": 5000,  "gdd_base_f": 50, "kc": 1.0},
}

CROP_MONITOR_URL = os.getenv(
    "CROP_MONITOR_URL",
    "https://oatmealfarmnetworkcropmonitorbackend-git-802455386518.us-central1.run.app"
    if os.getenv("GAE_ENV") or os.getenv("K_SERVICE")
    else "http://127.0.0.1:8002",
)

# ─── helpers ────────────────────────────────────────────────────────────────

def _field_or_404(field_id: int, db: Session):
    f = db.query(models.Field).filter(models.Field.FieldID == field_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Field not found")
    return f

def _safe_float(v):
    try:
        return float(v) if v is not None else None
    except Exception:
        return None

def _latest_analyses(field_id: int, limit: int = 50):
    try:
        r = requests.get(f"{CROP_MONITOR_URL}/api/fields/{field_id}/analyses?limit={limit}", timeout=10)
        return (r.json().get("analyses") or []) if r.ok else []
    except Exception:
        return []

def _get_index(analysis, name):
    for i in (analysis.get("vegetation_indices") or []):
        if (i.get("index_type") or "").upper() == name.upper():
            return i
    return None


# ═══════════════════════════════════════════════════════════════════
# SCOUTING
# ═══════════════════════════════════════════════════════════════════

def _ser_scout(row) -> dict:
    return {
        "scout_id":    row.ScoutID,
        "field_id":    row.FieldID,
        "business_id": row.BusinessID,
        "people_id":   row.PeopleID,
        "observed_at": row.ObservedAt.isoformat() + "Z" if row.ObservedAt else None,
        "category":    row.Category,
        "severity":    row.Severity,
        "notes":       row.Notes,
        "latitude":    float(row.Latitude) if row.Latitude is not None else None,
        "longitude":   float(row.Longitude) if row.Longitude is not None else None,
        "image_url":   row.ImageUrl,
        "created_at":  row.CreatedAt.isoformat() + "Z" if row.CreatedAt else None,
    }


@router.get("/fields/{field_id}/scouts")
def get_scouts(field_id: int, db: Session = Depends(get_db)):
    _field_or_404(field_id, db)
    rows = (
        db.query(models.FieldScout)
        .filter(models.FieldScout.FieldID == field_id)
        .order_by(desc(models.FieldScout.ObservedAt))
        .all()
    )
    return [_ser_scout(r) for r in rows]


@router.post("/fields/{field_id}/scouts")
def create_scout(field_id: int, body: dict, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    try:
        observed_at = datetime.fromisoformat(body["observed_at"].replace("Z", "")) if body.get("observed_at") else datetime.utcnow()
    except Exception:
        observed_at = datetime.utcnow()
    row = models.FieldScout(
        FieldID    = field_id,
        BusinessID = field.BusinessID,
        PeopleID   = body.get("people_id"),
        ObservedAt = observed_at,
        Category   = body.get("category", "General"),
        Severity   = body.get("severity"),
        Notes      = body.get("notes"),
        Latitude   = body.get("latitude"),
        Longitude  = body.get("longitude"),
        ImageUrl   = body.get("image_url"),
        CreatedAt  = datetime.utcnow(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _ser_scout(row)


@router.delete("/fields/{field_id}/scouts/{scout_id}")
def delete_scout(field_id: int, scout_id: int, db: Session = Depends(get_db)):
    row = db.query(models.FieldScout).filter(
        models.FieldScout.ScoutID == scout_id,
        models.FieldScout.FieldID == field_id,
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Scout observation not found")
    db.delete(row)
    db.commit()
    return {"success": True}


# ═══════════════════════════════════════════════════════════════════
# SOIL SAMPLES
# ═══════════════════════════════════════════════════════════════════

def _ser_soil(row) -> dict:
    return {
        "sample_id":     row.SampleID,
        "field_id":      row.FieldID,
        "business_id":   row.BusinessID,
        "sample_date":   str(row.SampleDate) if row.SampleDate else None,
        "sample_label":  row.SampleLabel,
        "latitude":      float(row.Latitude)      if row.Latitude      is not None else None,
        "longitude":     float(row.Longitude)     if row.Longitude     is not None else None,
        "depth_cm":      row.Depth_cm,
        "ph":            float(row.pH)            if row.pH            is not None else None,
        "organic_matter":float(row.OrganicMatter) if row.OrganicMatter is not None else None,
        "nitrogen":      float(row.Nitrogen)      if row.Nitrogen      is not None else None,
        "phosphorus":    float(row.Phosphorus)    if row.Phosphorus    is not None else None,
        "potassium":     float(row.Potassium)     if row.Potassium     is not None else None,
        "sulfur":        float(row.Sulfur)        if row.Sulfur        is not None else None,
        "calcium":       float(row.Calcium)       if row.Calcium       is not None else None,
        "magnesium":     float(row.Magnesium)     if row.Magnesium     is not None else None,
        "cec":           float(row.CEC)           if row.CEC           is not None else None,
        "notes":         row.Notes,
        "created_at":    row.CreatedAt.isoformat() + "Z" if row.CreatedAt else None,
    }


@router.get("/fields/{field_id}/soil-samples")
def get_soil_samples(field_id: int, db: Session = Depends(get_db)):
    _field_or_404(field_id, db)
    rows = (
        db.query(models.FieldSoilSample)
        .filter(models.FieldSoilSample.FieldID == field_id)
        .order_by(desc(models.FieldSoilSample.SampleDate))
        .all()
    )
    return [_ser_soil(r) for r in rows]


@router.post("/fields/{field_id}/soil-samples")
def create_soil_sample(field_id: int, body: dict, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    try:
        sample_date = date.fromisoformat(body["sample_date"]) if body.get("sample_date") else None
    except Exception:
        sample_date = None
    row = models.FieldSoilSample(
        FieldID       = field_id,
        BusinessID    = field.BusinessID,
        SampleDate    = sample_date,
        SampleLabel   = body.get("sample_label"),
        Latitude      = body.get("latitude"),
        Longitude     = body.get("longitude"),
        Depth_cm      = body.get("depth_cm"),
        pH            = body.get("ph"),
        OrganicMatter = body.get("organic_matter"),
        Nitrogen      = body.get("nitrogen"),
        Phosphorus    = body.get("phosphorus"),
        Potassium     = body.get("potassium"),
        Sulfur        = body.get("sulfur"),
        Calcium       = body.get("calcium"),
        Magnesium     = body.get("magnesium"),
        CEC           = body.get("cec"),
        Notes         = body.get("notes"),
        CreatedAt     = datetime.utcnow(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _ser_soil(row)


@router.delete("/fields/{field_id}/soil-samples/{sample_id}")
def delete_soil_sample(field_id: int, sample_id: int, db: Session = Depends(get_db)):
    row = db.query(models.FieldSoilSample).filter(
        models.FieldSoilSample.SampleID == sample_id,
        models.FieldSoilSample.FieldID  == field_id,
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Sample not found")
    db.delete(row)
    db.commit()
    return {"success": True}


@router.post("/fields/{field_id}/soil-samples/import")
async def import_soil_csv(field_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import soil samples from CSV. Expected columns (case-insensitive):
    sample_label, sample_date, latitude, longitude, depth_cm, ph, organic_matter,
    nitrogen, phosphorus, potassium, sulfur, calcium, magnesium, cec, notes
    """
    field = _field_or_404(field_id, db)
    content = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    rows_added = 0
    errors = []
    for i, row in enumerate(reader):
        r = {k.lower().strip(): v.strip() for k, v in row.items()}
        try:
            try:
                sd = date.fromisoformat(r.get("sample_date", "")) if r.get("sample_date") else None
            except Exception:
                sd = None
            db.add(models.FieldSoilSample(
                FieldID       = field_id,
                BusinessID    = field.BusinessID,
                SampleDate    = sd,
                SampleLabel   = r.get("sample_label") or r.get("label") or f"Sample {i+1}",
                Latitude      = _safe_float(r.get("latitude")),
                Longitude     = _safe_float(r.get("longitude")),
                Depth_cm      = int(r["depth_cm"]) if r.get("depth_cm") else None,
                pH            = _safe_float(r.get("ph")),
                OrganicMatter = _safe_float(r.get("organic_matter") or r.get("om")),
                Nitrogen      = _safe_float(r.get("nitrogen") or r.get("n")),
                Phosphorus    = _safe_float(r.get("phosphorus") or r.get("p")),
                Potassium     = _safe_float(r.get("potassium") or r.get("k")),
                Sulfur        = _safe_float(r.get("sulfur") or r.get("s")),
                Calcium       = _safe_float(r.get("calcium") or r.get("ca")),
                Magnesium     = _safe_float(r.get("magnesium") or r.get("mg")),
                CEC           = _safe_float(r.get("cec")),
                Notes         = r.get("notes"),
                CreatedAt     = datetime.utcnow(),
            ))
            rows_added += 1
        except Exception as e:
            errors.append(f"Row {i+2}: {e}")
    db.commit()
    return {"imported": rows_added, "errors": errors}


# ═══════════════════════════════════════════════════════════════════
# PRESCRIPTIONS
# ═══════════════════════════════════════════════════════════════════

def _ser_rx(row) -> dict:
    return {
        "prescription_id": row.PrescriptionID,
        "field_id":        row.FieldID,
        "business_id":     row.BusinessID,
        "name":            row.Name,
        "product":         row.Product,
        "unit":            row.Unit,
        "index_key":       row.IndexKey,
        "zone_method":     row.ZoneMethod,
        "num_zones":       row.NumZones,
        "zone_rates":      json.loads(row.ZoneRatesJSON) if row.ZoneRatesJSON else [],
        "analysis_date":   str(row.AnalysisDate) if row.AnalysisDate else None,
        "notes":           row.Notes,
        "created_at":      row.CreatedAt.isoformat() + "Z" if row.CreatedAt else None,
    }


@router.get("/fields/{field_id}/prescriptions")
def get_prescriptions(field_id: int, db: Session = Depends(get_db)):
    _field_or_404(field_id, db)
    rows = (
        db.query(models.FieldPrescription)
        .filter(models.FieldPrescription.FieldID == field_id)
        .order_by(desc(models.FieldPrescription.CreatedAt))
        .all()
    )
    return [_ser_rx(r) for r in rows]


@router.post("/fields/{field_id}/prescriptions")
def create_prescription(field_id: int, body: dict, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    try:
        analysis_date = date.fromisoformat(body["analysis_date"]) if body.get("analysis_date") else None
    except Exception:
        analysis_date = None
    row = models.FieldPrescription(
        FieldID       = field_id,
        BusinessID    = field.BusinessID,
        Name          = body.get("name", "Prescription"),
        Product       = body.get("product"),
        Unit          = body.get("unit"),
        IndexKey      = body.get("index_key", "NDVI"),
        ZoneMethod    = body.get("zone_method"),
        NumZones      = body.get("num_zones"),
        ZoneRatesJSON = json.dumps(body.get("zone_rates") or []),
        AnalysisDate  = analysis_date,
        Notes         = body.get("notes"),
        CreatedAt     = datetime.utcnow(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _ser_rx(row)


@router.delete("/fields/{field_id}/prescriptions/{rx_id}")
def delete_prescription(field_id: int, rx_id: int, db: Session = Depends(get_db)):
    row = db.query(models.FieldPrescription).filter(
        models.FieldPrescription.PrescriptionID == rx_id,
        models.FieldPrescription.FieldID        == field_id,
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Prescription not found")
    db.delete(row)
    db.commit()
    return {"success": True}


@router.get("/fields/{field_id}/prescriptions/{rx_id}/export.csv")
def export_prescription_csv(field_id: int, rx_id: int, db: Session = Depends(get_db)):
    row = db.query(models.FieldPrescription).filter(
        models.FieldPrescription.PrescriptionID == rx_id,
        models.FieldPrescription.FieldID        == field_id,
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Prescription not found")
    zone_rates = json.loads(row.ZoneRatesJSON) if row.ZoneRatesJSON else []
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Zone", "Rate", "Unit", "Product"])
    for zr in zone_rates:
        w.writerow([zr.get("zone"), zr.get("rate"), row.Unit or "", row.Product or ""])
    buf.seek(0)
    filename = f"prescription_{rx_id}_{row.Name or 'export'}.csv".replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════════
# WEATHER  (Open-Meteo — free, no key required)
# ═══════════════════════════════════════════════════════════════════

@router.get("/fields/{field_id}/wind")
def get_field_wind(field_id: int, days: int = 30, db: Session = Depends(get_db)):
    """Hourly wind direction + speed from Open-Meteo, aggregated into 8 compass
    sectors for a wind-rose plot. Drives spray-record drift documentation +
    disease-pressure modelling."""
    field = _field_or_404(field_id, db)
    lat = float(field.Latitude)  if field.Latitude  is not None else None
    lon = float(field.Longitude) if field.Longitude is not None else None
    if lat is None or lon is None:
        raise HTTPException(status_code=400, detail="Field has no coordinates")
    days = max(7, min(days, 90))

    try:
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={
                "latitude":         lat,
                "longitude":        lon,
                "hourly":           "wind_speed_10m,wind_direction_10m",
                "wind_speed_unit":  "kmh",
                "timezone":         "auto",
                "past_days":        days,
                "forecast_days":    1,
            },
            timeout=15,
        )
        r.raise_for_status()
        hourly = r.json().get("hourly", {})
    except requests.RequestException as e:
        raise HTTPException(status_code=502, detail=f"Weather service error: {e}")

    times = hourly.get("time", []) or []
    speeds = hourly.get("wind_speed_10m", []) or []
    dirs   = hourly.get("wind_direction_10m", []) or []

    # 8 compass sectors centered on N, NE, E, ..., NW. Each sector spans 45°.
    sector_labels = ["N","NE","E","SE","S","SW","W","NW"]
    sectors = [
        {"label": lbl, "count": 0, "speed_sum": 0.0, "speed_max": 0.0}
        for lbl in sector_labels
    ]
    speed_bins = [0, 5, 10, 15, 20, 30]   # km/h thresholds → 6 buckets
    bin_labels = ["calm <5", "5-10", "10-15", "15-20", "20-30", "30+"]
    matrix = [[0]*len(bin_labels) for _ in range(8)]   # [sector][bin] frequency

    total = 0
    for i in range(min(len(speeds), len(dirs), len(times))):
        sp  = speeds[i]
        dr  = dirs[i]
        if sp is None or dr is None:
            continue
        # Sector index — N spans -22.5°..22.5°, so shift by 22.5° before /45.
        sector_idx = int(((dr % 360) + 22.5) // 45) % 8
        s = sectors[sector_idx]
        s["count"]     += 1
        s["speed_sum"] += sp
        s["speed_max"]  = max(s["speed_max"], sp)
        # Speed bucket
        bin_idx = 0
        for b, thr in enumerate(speed_bins[1:], start=1):
            if sp >= thr:
                bin_idx = b
        matrix[sector_idx][bin_idx] += 1
        total += 1

    if total == 0:
        return {"field_id": field_id, "days": days, "samples": 0, "sectors": [], "matrix": [], "bin_labels": bin_labels}

    out_sectors = []
    for s in sectors:
        n = s["count"]
        out_sectors.append({
            "label":         s["label"],
            "count":         n,
            "frequency_pct": round(100.0 * n / total, 1),
            "mean_speed":    round(s["speed_sum"] / n, 1) if n else 0.0,
            "max_speed":     round(s["speed_max"], 1),
        })

    # Predominant direction = sector with highest count
    pred_idx = max(range(8), key=lambda i: sectors[i]["count"])
    return {
        "field_id":         field_id,
        "days":             days,
        "samples":          total,
        "predominant":      sector_labels[pred_idx],
        "predominant_pct":  round(100.0 * sectors[pred_idx]["count"] / total, 1),
        "calm_pct":         round(100.0 * sum(matrix[i][0] for i in range(8)) / total, 1),
        "sectors":          out_sectors,
        "matrix":           matrix,
        "bin_labels":       bin_labels,
        "speed_unit":       "kph",
    }


@router.get("/fields/{field_id}/weather")
def get_field_weather(field_id: int, days: int = 30, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    lat = float(field.Latitude)  if field.Latitude  is not None else None
    lon = float(field.Longitude) if field.Longitude is not None else None
    if lat is None or lon is None:
        raise HTTPException(status_code=400, detail="Field has no coordinates")
    days = max(7, min(days, 90))
    try:
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={
                "latitude":  lat,
                "longitude": lon,
                "daily": "temperature_2m_max,temperature_2m_min,precipitation_sum,et0_fao_evapotranspiration",
                "temperature_unit": "fahrenheit",
                "precipitation_unit": "inch",
                "timezone": "auto",
                "past_days": days,
                "forecast_days": 7,
            },
            timeout=10,
        )
        r.raise_for_status()
        data = r.json()
        daily = data.get("daily", {})
        dates     = daily.get("time", [])
        temp_max  = daily.get("temperature_2m_max", [])
        temp_min  = daily.get("temperature_2m_min", [])
        precip    = daily.get("precipitation_sum", [])
        et0       = daily.get("et0_fao_evapotranspiration", [])
        result = []
        for i, d in enumerate(dates):
            result.append({
                "date":     d,
                "temp_max": temp_max[i] if i < len(temp_max) else None,
                "temp_min": temp_min[i] if i < len(temp_min) else None,
                "precip":   precip[i]   if i < len(precip)   else None,
                "et0":      et0[i]      if i < len(et0)      else None,
            })
        return {"field_id": field_id, "lat": lat, "lon": lon, "daily": result}
    except requests.RequestException as e:
        raise HTTPException(status_code=502, detail=f"Weather service error: {e}")


# ═══════════════════════════════════════════════════════════════════
# REPORTS — Excel export
# ═══════════════════════════════════════════════════════════════════

@router.get("/fields/{field_id}/report.xlsx")
def export_field_report_xlsx(field_id: int, db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl not installed — run: pip install openpyxl")

    field = _field_or_404(field_id, db)
    analyses = _latest_analyses(field_id, limit=50)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Field info ──────────────────────────────────────
    ws = wb.active
    ws.title = "Field Info"
    header_fill = PatternFill("solid", fgColor="6D8E22")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(["Field Report — " + (field.Name or f"Field {field_id}")])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    ws.append([])
    ws.append(["Field Name",   field.Name])
    ws.append(["Crop Type",    field.CropType])
    ws.append(["Size (ha)",    float(field.FieldSizeHectares) if field.FieldSizeHectares else ""])
    ws.append(["Latitude",     float(field.Latitude)  if field.Latitude  else ""])
    ws.append(["Longitude",    float(field.Longitude) if field.Longitude else ""])
    ws.append(["Planting Date",str(field.PlantingDate) if field.PlantingDate else ""])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30

    # ── Sheet 2: Analysis history ────────────────────────────────
    if analyses:
        ws2 = wb.create_sheet("Analysis History")
        cols = ["Date", "Health Score", "NDVI", "NDRE", "EVI", "GNDVI", "NDWI", "Cloud %"]
        ws2.append(cols)
        for cell in ws2[1]:
            cell.fill  = header_fill
            cell.font  = header_font
            cell.alignment = Alignment(horizontal="center")
        for a in analyses:
            row = [
                (a.get("analysis_date") or "")[:10],
                a.get("health_score"),
            ]
            for idx in ["NDVI", "NDRE", "EVI", "GNDVI", "NDWI"]:
                d = _get_index(a, idx)
                row.append(round(d["mean"], 4) if d and d.get("mean") is not None else "")
            row.append(a.get("cloud_percent"))
            ws2.append(row)
        for i, _ in enumerate(cols, 1):
            ws2.column_dimensions[get_column_letter(i)].width = 14

    # ── Sheet 3: Soil samples ────────────────────────────────────
    soil_rows = (
        db.query(models.FieldSoilSample)
        .filter(models.FieldSoilSample.FieldID == field_id)
        .order_by(models.FieldSoilSample.SampleDate)
        .all()
    )
    if soil_rows:
        ws3 = wb.create_sheet("Soil Samples")
        soil_cols = ["Label", "Date", "Depth (cm)", "pH", "OM %", "N", "P", "K", "S", "Ca", "Mg", "CEC", "Notes"]
        ws3.append(soil_cols)
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
        for s in soil_rows:
            ws3.append([
                s.SampleLabel, str(s.SampleDate) if s.SampleDate else "",
                s.Depth_cm,
                float(s.pH)            if s.pH            is not None else "",
                float(s.OrganicMatter) if s.OrganicMatter is not None else "",
                float(s.Nitrogen)      if s.Nitrogen      is not None else "",
                float(s.Phosphorus)    if s.Phosphorus    is not None else "",
                float(s.Potassium)     if s.Potassium     is not None else "",
                float(s.Sulfur)        if s.Sulfur        is not None else "",
                float(s.Calcium)       if s.Calcium       is not None else "",
                float(s.Magnesium)     if s.Magnesium     is not None else "",
                float(s.CEC)           if s.CEC           is not None else "",
                s.Notes or "",
            ])
        for i, _ in enumerate(soil_cols, 1):
            ws3.column_dimensions[get_column_letter(i)].width = 12

    # ── Sheet 4: Scouts ──────────────────────────────────────────
    scout_rows = (
        db.query(models.FieldScout)
        .filter(models.FieldScout.FieldID == field_id)
        .order_by(desc(models.FieldScout.ObservedAt))
        .all()
    )
    if scout_rows:
        ws4 = wb.create_sheet("Scouting")
        scout_cols = ["Date", "Category", "Severity", "Notes", "Lat", "Lon"]
        ws4.append(scout_cols)
        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = header_font
        for s in scout_rows:
            ws4.append([
                s.ObservedAt.strftime("%Y-%m-%d") if s.ObservedAt else "",
                s.Category, s.Severity, s.Notes or "",
                float(s.Latitude)  if s.Latitude  is not None else "",
                float(s.Longitude) if s.Longitude is not None else "",
            ])
        for i, _ in enumerate(scout_cols, 1):
            ws4.column_dimensions[get_column_letter(i)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"field_{field_id}_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ═══════════════════════════════════════════════════════════════════
# ACTIVITY LOG
# ═══════════════════════════════════════════════════════════════════

def _ser_activity(row) -> dict:
    return {
        "activity_id":   row.ActivityID,
        "field_id":      row.FieldID,
        "business_id":   row.BusinessID,
        "people_id":     row.PeopleID,
        "activity_date": str(row.ActivityDate) if row.ActivityDate else None,
        "activity_type": row.ActivityType,
        "product":       row.Product,
        "rate":          float(row.Rate) if row.Rate is not None else None,
        "rate_unit":     row.RateUnit,
        "operator_name": row.OperatorName,
        "notes":         row.Notes,
        "created_at":    row.CreatedAt.isoformat() + "Z" if row.CreatedAt else None,
    }


@router.get("/fields/{field_id}/activity-log")
def get_activity_log(field_id: int, db: Session = Depends(get_db)):
    _field_or_404(field_id, db)
    rows = (
        db.query(models.FieldActivityLog)
        .filter(models.FieldActivityLog.FieldID == field_id)
        .order_by(desc(models.FieldActivityLog.ActivityDate))
        .all()
    )
    return [_ser_activity(r) for r in rows]


@router.post("/fields/{field_id}/activity-log")
def create_activity(field_id: int, body: dict, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    try:
        activity_date = date.fromisoformat(body["activity_date"]) if body.get("activity_date") else date.today()
    except Exception:
        activity_date = date.today()
    row = models.FieldActivityLog(
        FieldID      = field_id,
        BusinessID   = field.BusinessID,
        PeopleID     = body.get("people_id"),
        ActivityDate = activity_date,
        ActivityType = body.get("activity_type", "Other"),
        Product      = body.get("product"),
        Rate         = body.get("rate"),
        RateUnit     = body.get("rate_unit"),
        OperatorName = body.get("operator_name"),
        Notes        = body.get("notes"),
        CreatedAt    = datetime.utcnow(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _ser_activity(row)


@router.delete("/fields/{field_id}/activity-log/{activity_id}")
def delete_activity(field_id: int, activity_id: int, db: Session = Depends(get_db)):
    row = db.query(models.FieldActivityLog).filter(
        models.FieldActivityLog.ActivityID == activity_id,
        models.FieldActivityLog.FieldID   == field_id,
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Activity not found")
    db.delete(row)
    db.commit()
    return {"success": True}


# ═══════════════════════════════════════════════════════════════════
# FIELD BOUNDARY
# ═══════════════════════════════════════════════════════════════════

@router.get("/fields/{field_id}/boundary")
def get_boundary(field_id: int, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    geojson = None
    if field.BoundaryGeoJSON:
        try:
            geojson = json.loads(field.BoundaryGeoJSON)
        except Exception:
            geojson = None
    return {
        "field_id":   field_id,
        "name":       field.Name,
        "latitude":   float(field.Latitude)  if field.Latitude  is not None else None,
        "longitude":  float(field.Longitude) if field.Longitude is not None else None,
        "boundary":   geojson,
    }


@router.put("/fields/{field_id}/boundary")
def save_boundary(field_id: int, body: dict, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    geojson_obj = body.get("boundary")
    field.BoundaryGeoJSON = json.dumps(geojson_obj) if geojson_obj is not None else None
    # optionally update size from polygon area hint
    if body.get("field_size_hectares"):
        field.FieldSizeHectares = body["field_size_hectares"]
    db.commit()
    return {"success": True}


# ═══════════════════════════════════════════════════════════════════
# GDD — Growing Degree Days (uses Open-Meteo, same as weather)
# ═══════════════════════════════════════════════════════════════════

@router.get("/fields/{field_id}/gdd")
def get_gdd(
    field_id: int,
    base_temp_f: float = 50.0,
    days: int = 180,
    db: Session = Depends(get_db),
):
    field = _field_or_404(field_id, db)
    lat = float(field.Latitude)  if field.Latitude  is not None else None
    lon = float(field.Longitude) if field.Longitude is not None else None
    if lat is None or lon is None:
        raise HTTPException(status_code=400, detail="Field has no coordinates")

    # look up crop-type defaults
    crop_key = (field.CropType or "default").lower().split()[0]
    crop_info = _CROP_BASELINES.get(crop_key, _CROP_BASELINES["default"])
    effective_base = base_temp_f if base_temp_f != 50.0 else crop_info["gdd_base_f"]

    days = max(30, min(days, 365))
    # Open-Meteo's /v1/forecast caps past_days at 92, so use the historical
    # archive endpoint for windows beyond a week. The archive trails real-time
    # by a few days; that's fine for a GDD-accumulation chart.
    from datetime import date as _date, timedelta as _td
    end_date   = _date.today()
    start_date = end_date - _td(days=days)
    try:
        r = requests.get(
            "https://archive-api.open-meteo.com/v1/archive",
            params={
                "latitude":         lat,
                "longitude":        lon,
                "start_date":       start_date.isoformat(),
                "end_date":         end_date.isoformat(),
                "daily":            "temperature_2m_max,temperature_2m_min",
                "temperature_unit": "fahrenheit",
                "timezone":         "auto",
            },
            timeout=15,
        )
        if not r.ok:
            try:
                err_detail = r.json().get("reason") or r.text[:200]
            except Exception:
                err_detail = r.text[:200]
            raise HTTPException(status_code=502, detail=f"Weather archive error ({r.status_code}): {err_detail}")
        daily = r.json().get("daily", {}) or {}
    except requests.RequestException as e:
        raise HTTPException(status_code=502, detail=f"Weather service error: {e}")

    dates    = daily.get("time", [])
    temp_max = daily.get("temperature_2m_max", [])
    temp_min = daily.get("temperature_2m_min", [])

    result = []
    cumulative = 0.0
    for i, d in enumerate(dates):
        tmax = temp_max[i] if i < len(temp_max) and temp_max[i] is not None else effective_base
        tmin = temp_min[i] if i < len(temp_min) and temp_min[i] is not None else effective_base
        gdd = max(0.0, (tmax + tmin) / 2 - effective_base)
        cumulative += gdd
        result.append({"date": d, "gdd": round(gdd, 1), "cumulative": round(cumulative, 1)})

    return {
        "field_id":    field_id,
        "base_temp_f": effective_base,
        "crop_type":   field.CropType,
        "total_gdd":   round(cumulative, 1),
        "daily":       result,
    }


# ═══════════════════════════════════════════════════════════════════
# IRRIGATION SCHEDULING
# ═══════════════════════════════════════════════════════════════════

@router.get("/fields/{field_id}/irrigation")
def get_irrigation(field_id: int, days: int = 30, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    lat = float(field.Latitude)  if field.Latitude  is not None else None
    lon = float(field.Longitude) if field.Longitude is not None else None
    if lat is None or lon is None:
        raise HTTPException(status_code=400, detail="Field has no coordinates")

    crop_key = (field.CropType or "default").lower().split()[0]
    kc = _CROP_BASELINES.get(crop_key, _CROP_BASELINES["default"])["kc"]
    days = max(7, min(days, 60))

    try:
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={
                "latitude":  lat,
                "longitude": lon,
                "daily": "precipitation_sum,et0_fao_evapotranspiration",
                "precipitation_unit": "inch",
                "timezone": "auto",
                "past_days": days,
                "forecast_days": 7,
            },
            timeout=10,
        )
        r.raise_for_status()
        daily = r.json().get("daily", {})
    except requests.RequestException as e:
        raise HTTPException(status_code=502, detail=f"Weather service error: {e}")

    dates  = daily.get("time", [])
    precip = daily.get("precipitation_sum", [])
    et0    = daily.get("et0_fao_evapotranspiration", [])

    result = []
    cumulative_deficit = 0.0
    for i, d in enumerate(dates):
        p   = precip[i] if i < len(precip) and precip[i] is not None else 0.0
        e0  = et0[i]    if i < len(et0)    and et0[i]    is not None else 0.0
        etc = e0 * kc   # crop evapotranspiration
        deficit = max(0.0, etc - p)
        surplus = max(0.0, p - etc)
        cumulative_deficit = max(0.0, cumulative_deficit + deficit - surplus)
        result.append({
            "date": d,
            "precip_in": round(p, 3),
            "et0_in": round(e0, 3),
            "etc_in": round(etc, 3),
            "deficit_in": round(deficit, 3),
            "cumulative_deficit_in": round(cumulative_deficit, 3),
        })

    # recommendation based on last 7 days cumulative deficit
    recent_deficit = result[-7]["cumulative_deficit_in"] if len(result) >= 7 else cumulative_deficit
    if recent_deficit >= 1.5:
        recommendation = "Irrigate now"
        urgency = "high"
    elif recent_deficit >= 0.75:
        recommendation = "Consider irrigating within 2–3 days"
        urgency = "medium"
    else:
        recommendation = "No irrigation needed"
        urgency = "low"

    return {
        "field_id":          field_id,
        "crop_type":         field.CropType,
        "kc":                kc,
        "recommendation":    recommendation,
        "urgency":           urgency,
        "cumulative_deficit_in": round(cumulative_deficit, 3),
        "daily":             result,
    }


# ═══════════════════════════════════════════════════════════════════
# YIELD FORECAST
# ═══════════════════════════════════════════════════════════════════

@router.get("/fields/{field_id}/yield-forecast")
def get_yield_forecast(field_id: int, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    analyses = _latest_analyses(field_id, limit=20)

    crop_key = (field.CropType or "default").lower().split()[0]
    baseline = _CROP_BASELINES.get(crop_key, _CROP_BASELINES["default"])["yield_kgha"]

    if not analyses:
        return {
            "field_id": field_id,
            "crop_type": field.CropType,
            "forecast_kgha": None,
            "confidence": "low",
            "message": "No satellite analyses available",
            "history": [],
        }

    history = []
    for a in analyses[:10]:
        ndvi_d = _get_index(a, "NDVI")
        ndvi   = ndvi_d["mean"] if ndvi_d and ndvi_d.get("mean") is not None else None
        if ndvi is not None:
            # NDVI-to-yield: linear scale against optimal NDVI of 0.75
            estimated = max(0, baseline * (ndvi / 0.75))
            history.append({
                "date":         (a.get("analysis_date") or "")[:10],
                "ndvi":         round(ndvi, 4),
                "forecast_kgha": round(estimated),
                "health_score": a.get("health_score"),
            })

    latest = history[0] if history else None

    # trend — compare first and last if we have enough data
    confidence = "low"
    trend_pct = None
    if len(history) >= 3:
        old_yield = history[-1]["forecast_kgha"]
        new_yield = history[0]["forecast_kgha"]
        if old_yield > 0:
            trend_pct = round((new_yield - old_yield) / old_yield * 100, 1)
        confidence = "medium" if len(history) >= 5 else "low"
    if len(history) >= 8:
        confidence = "high"

    return {
        "field_id":      field_id,
        "crop_type":     field.CropType,
        "baseline_kgha": baseline,
        "forecast_kgha": latest["forecast_kgha"] if latest else None,
        "confidence":    confidence,
        "trend_pct":     trend_pct,
        "history":       history,
    }


# ═══════════════════════════════════════════════════════════════════
# ALERTS  (generated from own data + crop monitoring backend)
# ═══════════════════════════════════════════════════════════════════

@router.get("/fields/{field_id}/alerts")
def get_field_alerts(field_id: int, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)
    alerts = []

    # 1. Low health score from latest analysis
    analyses = _latest_analyses(field_id, limit=3)
    if analyses:
        latest = analyses[0]
        score = latest.get("health_score")
        if score is not None and score < 50:
            alerts.append({
                "alert_id":  f"health_{field_id}",
                "type":      "Health",
                "severity":  "Critical" if score < 30 else "High",
                "message":   f"Field health score is {score}% — below acceptable threshold",
                "date":      (latest.get("analysis_date") or "")[:10],
                "source":    "satellite",
                "acknowledged": False,
            })
        elif score is not None and score < 70:
            alerts.append({
                "alert_id":  f"health_warn_{field_id}",
                "type":      "Health",
                "severity":  "Medium",
                "message":   f"Field health score is {score}% — monitoring recommended",
                "date":      (latest.get("analysis_date") or "")[:10],
                "source":    "satellite",
                "acknowledged": False,
            })

        # NDVI decline alert
        if len(analyses) >= 2:
            ndvi_now  = (_get_index(analyses[0], "NDVI") or {}).get("mean")
            ndvi_prev = (_get_index(analyses[1], "NDVI") or {}).get("mean")
            if ndvi_now is not None and ndvi_prev is not None:
                drop = ndvi_prev - ndvi_now
                if drop > 0.1:
                    alerts.append({
                        "alert_id":  f"ndvi_drop_{field_id}",
                        "type":      "NDVI Decline",
                        "severity":  "High" if drop > 0.2 else "Medium",
                        "message":   f"NDVI dropped by {drop:.3f} since last analysis",
                        "date":      (analyses[0].get("analysis_date") or "")[:10],
                        "source":    "satellite",
                        "acknowledged": False,
                    })

    # 2. High-severity scouting observations in last 14 days
    cutoff = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    from datetime import timedelta
    cutoff14 = cutoff - timedelta(days=14)
    recent_scouts = (
        db.query(models.FieldScout)
        .filter(
            models.FieldScout.FieldID == field_id,
            models.FieldScout.ObservedAt >= cutoff14,
            models.FieldScout.Severity.in_(["High", "Critical"]),
        )
        .order_by(desc(models.FieldScout.ObservedAt))
        .limit(5)
        .all()
    )
    for s in recent_scouts:
        alerts.append({
            "alert_id":  f"scout_{s.ScoutID}",
            "type":      s.Category or "Scouting",
            "severity":  s.Severity,
            "message":   s.Notes or f"{s.Severity} severity {s.Category} observation",
            "date":      s.ObservedAt.strftime("%Y-%m-%d") if s.ObservedAt else None,
            "source":    "scouting",
            "acknowledged": False,
        })

    # 3. Try crop monitoring backend alerts
    try:
        r = requests.get(f"{CROP_MONITOR_URL}/api/fields/{field_id}/alerts", timeout=5)
        if r.ok:
            for a in (r.json() or []):
                a["source"] = "crop_monitor"
                alerts.append(a)
    except Exception:
        pass

    return {"field_id": field_id, "alerts": alerts}


# ═══════════════════════════════════════════════════════════════════
# BENCHMARK — compare all fields for a business
# ═══════════════════════════════════════════════════════════════════

@router.get("/businesses/{business_id}/benchmark")
def get_benchmark(business_id: int, db: Session = Depends(get_db)):
    fields = (
        db.query(models.Field)
        .filter(models.Field.BusinessID == business_id, models.Field.DeletedAt.is_(None))
        .all()
    )
    if not fields:
        return {"business_id": business_id, "fields": []}

    results = []
    for field in fields:
        analyses = _latest_analyses(field.FieldID, limit=10)
        if not analyses:
            results.append({
                "field_id":   field.FieldID,
                "name":       field.Name,
                "crop_type":  field.CropType,
                "ndvi":       None,
                "health":     None,
                "trend":      None,
                "analyses":   0,
            })
            continue

        latest      = analyses[0]
        ndvi_now    = (_get_index(latest, "NDVI") or {}).get("mean")
        ndvi_prev   = (_get_index(analyses[1], "NDVI") or {}).get("mean") if len(analyses) > 1 else None
        trend       = round(ndvi_now - ndvi_prev, 4) if ndvi_now is not None and ndvi_prev is not None else None

        results.append({
            "field_id":   field.FieldID,
            "name":       field.Name,
            "crop_type":  field.CropType,
            "ndvi":       round(ndvi_now, 4) if ndvi_now is not None else None,
            "health":     latest.get("health_score"),
            "trend":      trend,
            "analyses":   len(analyses),
            "last_date":  (latest.get("analysis_date") or "")[:10],
        })

    # sort by NDVI descending (None last)
    results.sort(key=lambda x: (x["ndvi"] is None, -(x["ndvi"] or 0)))
    return {"business_id": business_id, "fields": results}


# ═══════════════════════════════════════════════════════════════════
# CARBON / SUSTAINABILITY  — soil OM trends + cover crop history
# ═══════════════════════════════════════════════════════════════════

@router.get("/fields/{field_id}/carbon")
def get_carbon(field_id: int, db: Session = Depends(get_db)):
    field = _field_or_404(field_id, db)

    # Soil OM samples over time
    soil_rows = (
        db.query(models.FieldSoilSample)
        .filter(models.FieldSoilSample.FieldID == field_id)
        .order_by(models.FieldSoilSample.SampleDate)
        .all()
    )
    om_history = []
    for s in soil_rows:
        if s.OrganicMatter is not None:
            om_pct = float(s.OrganicMatter)
            # Approximate soil carbon: OM% × 0.58 = SOC%
            # SOC stock (Mg C/ha) = SOC% / 100 × bulk density (1.3 t/m³) × depth (cm) × 100
            depth = s.Depth_cm or 30
            soc_stock = (om_pct * 0.58 / 100) * 1.3 * depth * 100  # Mg C/ha
            om_history.append({
                "date":       str(s.SampleDate) if s.SampleDate else None,
                "label":      s.SampleLabel,
                "om_pct":     om_pct,
                "soc_pct":    round(om_pct * 0.58, 2),
                "soc_stock_MgCha": round(soc_stock, 1),
            })

    # Crop rotation history
    rotations = (
        db.query(models.CropRotationEntry)
        .filter(models.CropRotationEntry.FieldID == field_id)
        .order_by(desc(models.CropRotationEntry.SeasonYear))
        .all()
    )
    rotation_history = []
    cover_crop_seasons = 0
    for rot in rotations:
        rotation_history.append({
            "season_year":   rot.SeasonYear,
            "crop":          rot.CropName,
            "variety":       rot.Variety,
            "planting_date": str(rot.PlantingDate) if rot.PlantingDate else None,
            "harvest_date":  str(rot.HarvestDate)  if rot.HarvestDate  else None,
            "yield_amount":  float(rot.YieldAmount) if rot.YieldAmount is not None else None,
            "yield_unit":    rot.YieldUnit,
            "is_cover_crop": rot.IsCoverCrop,
            "notes":         rot.Notes,
        })
        if rot.IsCoverCrop:
            cover_crop_seasons += 1

    # Carbon trend (first vs last OM sample)
    om_trend = None
    if len(om_history) >= 2:
        om_trend = round(om_history[-1]["om_pct"] - om_history[0]["om_pct"], 2)

    # Sustainability score (0-100)
    score = 50
    if cover_crop_seasons > 0: score += min(20, cover_crop_seasons * 5)
    if om_trend is not None and om_trend > 0: score += min(20, int(om_trend * 20))
    if len(rotations) >= 3:
        unique_crops = len(set(r.CropName for r in rotations if r.CropName))
        if unique_crops >= 3: score += 10
    score = min(100, score)

    return {
        "field_id":            field_id,
        "crop_type":           field.CropType,
        "sustainability_score": score,
        "cover_crop_seasons":  cover_crop_seasons,
        "om_trend_pct":        om_trend,
        "latest_soc_MgCha":    om_history[-1]["soc_stock_MgCha"] if om_history else None,
        "om_history":          om_history,
        "rotation_history":    rotation_history,
    }
